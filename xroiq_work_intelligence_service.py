import json
import logging
import os
import queue
import shutil
import sys
import threading
import time
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

from openpyxl import load_workbook
from watchdog.events import FileSystemEventHandler
from watchdog.observers import Observer

from xroiq_store import init_db, insert_event, insert_session


APP_NAME = "XROIQ Work Intelligence Service"
DEFAULT_CONFIG_PATH = Path(__file__).with_name("xroiq_work_intelligence_config.json")
DEFAULT_DATABASE_PATH = r"C:\\XROIQ\\ops\\data\\xroiq_work_intelligence.db"


DEFAULT_CONFIG = {
    "watch_paths": [r"C:\\XROIQ\\work"],
    "ignore_dirs": [".git", "node_modules", ".next", "dist", "build", "__pycache__", ".venv", "venv", ".idea", ".vs", ".turbo"],
    "ignore_extensions": [".tmp", ".log", ".cache", ".ds_store", ".part", ".crdownload", ".lock"],
    "session_gap_minutes": 10,
    "event_debounce_seconds": 8,
    "heartbeat_interval_seconds": 300,
    "shutdown_drain_timeout_seconds": 10,
    "queue_warning_size": 500,
    "owner": "Karne",
    "device": "LAP-002",
    "laptop_root": r"C:\\XROIQ\\work",
    "microsd_root": r"E:\\XROIQ_STAGE",
    "wd_root": r"D:\\XROIQ_ARCHIVE",
    "copy_recent_to_microsd": True,
    "copy_stable_to_wd": True,
    "stable_days_before_wd": 7,
    "workbook_path": r"C:\\XROIQ\\ops\\XROIQ_Work_Intelligence_Template.xlsx",
    "backup_root": r"C:\\XROIQ\\ops\\logs",
    "database_path": DEFAULT_DATABASE_PATH,
    "category_rules": {
        "Build": {
            "extensions": [".ts", ".tsx", ".js", ".jsx", ".py", ".json", ".yml", ".yaml", ".env", ".sql", ".sh", ".ps1", ".bat"]
        },
        "R&D": {
            "extensions": [".md", ".txt", ".pdf", ".csv", ".parquet", ".ipynb"]
        },
        "Admin": {
            "extensions": [".docx", ".xlsx", ".pptx", ".msg", ".rtf"]
        },
        "Communications": {
            "path_keywords": ["mail", "outlook", "inbox", "teams", "slack", "meeting", "calendar"]
        }
    },
    "importance_keywords": {
        "High": ["final", "production", "invoice", "contract", "launch", "master"],
        "Medium": ["draft", "review", "notes"],
        "Low": []
    }
}

REQUIRED_CONFIG_FIELDS = [
    "watch_paths",
    "ignore_dirs",
    "ignore_extensions",
    "session_gap_minutes",
    "event_debounce_seconds",
    "owner",
    "device",
    "laptop_root",
    "workbook_path",
    "backup_root",
    "category_rules",
    "importance_keywords",
]

OPTIONAL_CONFIG_DEFAULTS = {
    "heartbeat_interval_seconds": 300,
    "shutdown_drain_timeout_seconds": 10,
    "queue_warning_size": 500,
}

COUNTER_NAMES = [
    "events_enqueued",
    "events_ignored",
    "events_processed",
    "events_failed",
    "fallback_events_written",
    "fallback_events_failed",
    "copy_success",
    "copy_failed",
    "copy_skipped",
    "queue_warning_count",
]

REQUIRED_ACTIVITY_HEADERS = [
    "Event_ID",
    "Event_Time",
    "Event_Type",
    "File_Name",
    "Full_Path",
    "Extension",
    "Parent_Folder",
    "Category",
    "Importance",
    "Device",
    "Source",
    "Session_Key",
    "Handled_Action",
    "Backup_Status",
    "Notes",
]

REQUIRED_SESSION_HEADERS = [
    "Session_Key",
    "Date",
    "Start_Time",
    "End_Time",
    "Duration_Min",
    "Category",
    "Primary_Path",
    "File_Count",
    "Owner",
]


class ConfigurationError(RuntimeError):
    """Raised when the service configuration is invalid."""


class WorkbookValidationError(RuntimeError):
    """Raised when the workbook structure is invalid."""


@dataclass
class EventRecord:
    event_type: str
    path: Path
    when: datetime
    source_path: Optional[Path] = None
    dest_path: Optional[Path] = None


class DebouncedHandler(FileSystemEventHandler):
    def __init__(self, service: "WorkIntelligenceService"):
        self.service = service

    def on_created(self, event):
        self.service.enqueue_event("created", event)

    def on_modified(self, event):
        self.service.enqueue_event("modified", event)

    def on_moved(self, event):
        self.service.enqueue_event("moved", event)

    def on_deleted(self, event):
        self.service.enqueue_event("deleted", event)


class WorkbookLogger:
    def __init__(self, workbook_path: Path):
        self.workbook_path = workbook_path
        self.lock = threading.Lock()

    def _ensure_parent(self) -> None:
        self.workbook_path.parent.mkdir(parents=True, exist_ok=True)

    def append_activity(self, row: List):
        with self.lock:
            self._ensure_parent()
            wb = None
            try:
                wb = load_workbook(self.workbook_path)
                ws = wb["Activity_Log"]
                ws.append(row)
                wb.save(self.workbook_path)
            finally:
                if wb is not None:
                    wb.close()

    def append_or_extend_session(self, row: Dict, gap_minutes: int):
        with self.lock:
            self._ensure_parent()
            wb = None
            try:
                wb = load_workbook(self.workbook_path)
                ws = wb["Sessions"]
                last_row = ws.max_row
                now_start = row["start_time"]
                now_end = row["end_time"]
                category = row["category"]
                path_root = row["primary_path"]
                owner = row["owner"]

                if last_row >= 2:
                    last_category = ws[f"F{last_row}"].value
                    last_end = ws[f"D{last_row}"].value
                    if isinstance(last_end, datetime) and last_category == category and (now_start - last_end) <= timedelta(minutes=gap_minutes):
                        ws[f"D{last_row}"] = now_end
                        start_time = ws[f"C{last_row}"].value
                        if isinstance(start_time, datetime):
                            minutes = max(1, int((now_end - start_time).total_seconds() // 60))
                            ws[f"E{last_row}"] = minutes
                        ws[f"G{last_row}"] = path_root
                        current_count = ws[f"H{last_row}"].value or 0
                        ws[f"H{last_row}"] = int(current_count) + 1
                        wb.save(self.workbook_path)
                        return ws[f"A{last_row}"].value

                session_key = f"{now_start.strftime('%Y%m%d-%H%M')}-{category}"
                ws.append([
                    session_key,
                    now_start.date(),
                    now_start,
                    now_end,
                    1,
                    category,
                    path_root,
                    1,
                    owner,
                ])
                wb.save(self.workbook_path)
                return session_key
            finally:
                if wb is not None:
                    wb.close()

    def validate_workbook_structure(self) -> None:
        if not self.workbook_path.exists():
            raise WorkbookValidationError(
                f"Workbook file does not exist: {self.workbook_path}"
            )

        wb = load_workbook(self.workbook_path)
        try:
            self._validate_sheet_headers(
                workbook=wb,
                sheet_name="Activity_Log",
                required_headers=REQUIRED_ACTIVITY_HEADERS,
            )
            self._validate_sheet_headers(
                workbook=wb,
                sheet_name="Sessions",
                required_headers=REQUIRED_SESSION_HEADERS,
            )
        finally:
            wb.close()

    @staticmethod
    def _validate_sheet_headers(
        workbook,
        sheet_name: str,
        required_headers: Sequence[str],
    ) -> None:
        if sheet_name not in workbook.sheetnames:
            raise WorkbookValidationError(
                f"Workbook is missing required sheet: {sheet_name}"
            )

        worksheet = workbook[sheet_name]
        headers = [
            "" if cell.value is None else str(cell.value).strip()
            for cell in worksheet[1]
        ]
        missing_headers = [header for header in required_headers if header not in headers]
        if missing_headers:
            raise WorkbookValidationError(
                f"Sheet '{sheet_name}' is missing required headers: {', '.join(missing_headers)}"
            )


class FallbackEventWriter:
    def __init__(self, path: Path):
        self.path = path
        self.lock = threading.Lock()

    def append_record(self, record: Dict[str, Any]) -> None:
        with self.lock:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            line = json.dumps(record, ensure_ascii=True) + "\n"
            with open(self.path, "a", encoding="utf-8") as fallback_file:
                fallback_file.write(line)


class WorkIntelligenceService:
    def __init__(self, config_path: Path, config_path_was_explicit: bool = False):
        self.config_path = config_path
        loaded_config = self.load_config(
            config_path,
            allow_create_default=not config_path_was_explicit,
        )
        self.config = self.apply_optional_defaults(loaded_config)
        self.validated_paths = self.validate_config(self.config)
        self.logger = WorkbookLogger(Path(self.config["workbook_path"]))
        self.database_path = self.validated_paths["database_path"]
        self.fallback_writer = FallbackEventWriter(
            self.validated_paths["backup_root"] / "fallback_events.jsonl"
        )
        self.event_queue: "queue.Queue[EventRecord]" = queue.Queue()
        self.stop_event = threading.Event()
        self.stop_lock = threading.Lock()
        self.shutdown_started = False
        self.observer = Observer()
        self.observer_started = False
        self.worker_thread: Optional[threading.Thread] = None
        self.last_seen: Dict[str, datetime] = {}
        self.event_counter = 1
        self.counter_lock = threading.Lock()
        self.counters = self._create_counters()
        self.last_heartbeat_at = datetime.now()
        self.log = self._create_logger()

    @staticmethod
    def load_config(config_path: Path, allow_create_default: bool) -> Dict:
        if not config_path.exists():
            if not allow_create_default:
                raise ConfigurationError(
                    f"Config file does not exist: {config_path}"
                )
            config_path.parent.mkdir(parents=True, exist_ok=True)
            config_path.write_text(json.dumps(DEFAULT_CONFIG, indent=2), encoding="utf-8")
            return DEFAULT_CONFIG
        with open(config_path, "r", encoding="utf-8") as f:
            return json.load(f)

    @staticmethod
    def apply_optional_defaults(config: Dict) -> Dict:
        merged_config = dict(config)
        for field, default_value in OPTIONAL_CONFIG_DEFAULTS.items():
            merged_config.setdefault(field, default_value)
        merged_config.setdefault("database_path", DEFAULT_DATABASE_PATH)
        return merged_config

    @staticmethod
    def validate_config(config: Dict) -> Dict[str, Path]:
        missing_fields = [field for field in REQUIRED_CONFIG_FIELDS if field not in config]
        if missing_fields:
            raise ConfigurationError(
                f"Config is missing required field(s): {', '.join(missing_fields)}"
            )

        if not isinstance(config["watch_paths"], list) or not config["watch_paths"]:
            raise ConfigurationError("Config field 'watch_paths' must be a non-empty list.")
        if not isinstance(config["ignore_dirs"], list):
            raise ConfigurationError("Config field 'ignore_dirs' must be a list.")
        if not isinstance(config["ignore_extensions"], list):
            raise ConfigurationError("Config field 'ignore_extensions' must be a list.")
        if not isinstance(config["category_rules"], dict) or not config["category_rules"]:
            raise ConfigurationError("Config field 'category_rules' must be a non-empty object.")
        if not isinstance(config["importance_keywords"], dict):
            raise ConfigurationError("Config field 'importance_keywords' must be an object.")

        for field in ("owner", "device", "laptop_root", "workbook_path", "backup_root"):
            if not isinstance(config[field], str) or not config[field].strip():
                raise ConfigurationError(f"Config field '{field}' must be a non-empty string.")

        session_gap = config["session_gap_minutes"]
        if not isinstance(session_gap, int) or session_gap <= 0:
            raise ConfigurationError(
                "Config field 'session_gap_minutes' must be a positive integer."
            )

        debounce_seconds = config["event_debounce_seconds"]
        if not isinstance(debounce_seconds, int) or debounce_seconds < 0:
            raise ConfigurationError(
                "Config field 'event_debounce_seconds' must be an integer >= 0."
            )

        for field in OPTIONAL_CONFIG_DEFAULTS:
            value = config.get(field, OPTIONAL_CONFIG_DEFAULTS[field])
            if not isinstance(value, int) or value <= 0:
                raise ConfigurationError(
                    f"Config field '{field}' must be a positive integer."
                )

        if not isinstance(config.get("database_path"), str) or not config["database_path"].strip():
            raise ConfigurationError("Config field 'database_path' must be a non-empty string.")

        validated_paths = {
            "laptop_root": Path(config["laptop_root"]).expanduser(),
            "workbook_path": Path(config["workbook_path"]).expanduser(),
            "backup_root": Path(config["backup_root"]).expanduser(),
            "database_path": Path(config["database_path"]).expanduser(),
            "microsd_root": Path(config["microsd_root"]).expanduser()
            if config.get("microsd_root")
            else None,
            "wd_root": Path(config["wd_root"]).expanduser()
            if config.get("wd_root")
            else None,
        }
        watch_paths = []
        for watch_path in config["watch_paths"]:
            if not isinstance(watch_path, str) or not watch_path.strip():
                raise ConfigurationError(
                    "Each value in 'watch_paths' must be a non-empty string."
                )
            watch_paths.append(Path(watch_path).expanduser())
        validated_paths["watch_paths"] = watch_paths

        backup_root = validated_paths["backup_root"]
        try:
            backup_root.mkdir(parents=True, exist_ok=True)
        except OSError as exc:
            raise ConfigurationError(
                f"Backup root is not writable or cannot be created: {backup_root} ({exc})"
            ) from exc

        workbook_path = validated_paths["workbook_path"]
        if not workbook_path.exists():
            raise ConfigurationError(f"Workbook path does not exist: {workbook_path}")
        if workbook_path.suffix.lower() != ".xlsx":
            raise ConfigurationError("Config field 'workbook_path' must point to an .xlsx file.")

        laptop_root = validated_paths["laptop_root"]
        if not laptop_root.exists():
            raise ConfigurationError(
                f"Laptop root does not exist. Create it before starting the service: {laptop_root}"
            )
        if not laptop_root.is_dir():
            raise ConfigurationError(f"Laptop root must be a directory: {laptop_root}")

        for watch_path in watch_paths:
            if not watch_path.exists():
                raise ConfigurationError(
                    f"Watch path does not exist. Create it before starting the service: {watch_path}"
                )
            if not watch_path.is_dir():
                raise ConfigurationError(f"Watch path must be a directory: {watch_path}")

        for category, rule in config["category_rules"].items():
            if not isinstance(rule, dict):
                raise ConfigurationError(
                    f"Category rule '{category}' must be an object."
                )
            extensions = rule.get("extensions", [])
            path_keywords = rule.get("path_keywords", [])
            if extensions and not isinstance(extensions, list):
                raise ConfigurationError(
                    f"Category rule '{category}.extensions' must be a list."
                )
            if path_keywords and not isinstance(path_keywords, list):
                raise ConfigurationError(
                    f"Category rule '{category}.path_keywords' must be a list."
                )

        for level, keywords in config["importance_keywords"].items():
            if not isinstance(keywords, list):
                raise ConfigurationError(
                    f"Importance rule '{level}' must be a list of keywords."
                )

        return validated_paths

    def _create_logger(self) -> logging.Logger:
        log = logging.getLogger(APP_NAME)
        if log.handlers:
            return log
        log.setLevel(logging.INFO)
        formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
        log_dir = Path(self.config["backup_root"])
        log_dir.mkdir(parents=True, exist_ok=True)
        fh = logging.FileHandler(log_dir / "work_intelligence_service.log", encoding="utf-8")
        fh.setFormatter(formatter)
        log.addHandler(fh)
        sh = logging.StreamHandler(sys.stdout)
        sh.setFormatter(formatter)
        log.addHandler(sh)
        return log

    @staticmethod
    def _create_counters() -> Dict[str, int]:
        return {name: 0 for name in COUNTER_NAMES}

    def _increment_counter(self, name: str, amount: int = 1) -> None:
        with self.counter_lock:
            self.counters[name] = self.counters.get(name, 0) + amount

    def _counter_snapshot(self) -> Dict[str, int]:
        with self.counter_lock:
            return dict(self.counters)

    def _format_counter_summary(self) -> str:
        snapshot = self._counter_snapshot()
        return ", ".join(f"{name}={snapshot.get(name, 0)}" for name in COUNTER_NAMES)

    def _log_heartbeat_if_due(self, *, force: bool = False) -> None:
        now = datetime.now()
        interval = int(self.config.get("heartbeat_interval_seconds", 300))
        if not force and (now - self.last_heartbeat_at).total_seconds() < interval:
            return
        self.last_heartbeat_at = now
        self.log.info(
            "Heartbeat | queue_size=%s | %s",
            self.event_queue.qsize(),
            self._format_counter_summary(),
        )

    def _wait_for_queue_drain(self, timeout_seconds: int) -> bool:
        deadline = time.monotonic() + timeout_seconds
        while self.event_queue.unfinished_tasks > 0 and time.monotonic() < deadline:
            time.sleep(0.1)
        return self.event_queue.unfinished_tasks == 0

    def validate_runtime(self) -> None:
        self.logger.validate_workbook_structure()
        try:
            init_db(self.database_path)
            self.log.info("SQLite database initialized: %s", self.database_path)
        except Exception:
            self.log.error(
                "SQLite database initialization failed; continuing with Excel logging",
                exc_info=True,
            )
        self._warn_optional_mounts()
        self._log_startup_summary()

    def _warn_optional_mounts(self) -> None:
        microsd_root = self.validated_paths.get("microsd_root")
        if microsd_root and not microsd_root.exists():
            self.log.warning(
                "Optional microSD root is not mounted or missing: %s",
                microsd_root,
            )

        wd_root = self.validated_paths.get("wd_root")
        if wd_root and not wd_root.exists():
            self.log.warning(
                "Optional WD archive root is not mounted or missing: %s",
                wd_root,
            )

    def _log_startup_summary(self) -> None:
        self.log.info("Validated config: watch_paths=%s", ", ".join(str(path) for path in self.validated_paths["watch_paths"]))
        self.log.info("Validated workbook: %s", self.validated_paths["workbook_path"])
        self.log.info("Validated laptop root: %s", self.validated_paths["laptop_root"])
        self.log.info("Validated backup root: %s", self.validated_paths["backup_root"])
        self.log.info("Validated SQLite database path: %s", self.database_path)
        self.log.info("Fallback event log: %s", self.fallback_writer.path)

    def should_ignore(self, path: Path) -> bool:
        lower_parts = [p.lower() for p in path.parts]
        ignore_dirs = {d.lower() for d in self.config.get("ignore_dirs", [])}
        if any(part in ignore_dirs for part in lower_parts):
            return True
        if path.suffix.lower() in {e.lower() for e in self.config.get("ignore_extensions", [])}:
            return True
        return False

    def enqueue_event(self, event_type: str, event) -> None:
        if event.is_directory:
            self._increment_counter("events_ignored")
            return

        source_path = Path(event.src_path) if getattr(event, "src_path", None) else None
        dest_path_value = getattr(event, "dest_path", None)
        dest_path = Path(dest_path_value) if dest_path_value else None
        path = dest_path or source_path
        if path is None:
            self._increment_counter("events_ignored")
            return

        if event_type == "moved":
            ignored_paths = [
                self.should_ignore(candidate)
                for candidate in (source_path, dest_path)
                if candidate is not None
            ]
            if ignored_paths and all(ignored_paths):
                self._increment_counter("events_ignored")
                return
        elif self.should_ignore(path):
            self._increment_counter("events_ignored")
            return

        now = datetime.now()
        if event_type == "moved":
            key = f"{event_type}|{str(source_path).lower()}|{str(dest_path).lower()}"
        else:
            key = f"{event_type}|{str(path).lower()}"
        debounce = int(self.config.get("event_debounce_seconds", 8))
        last = self.last_seen.get(key)
        if last and (now - last).total_seconds() < debounce:
            self._increment_counter("events_ignored")
            return
        self.last_seen[key] = now
        self.event_queue.put(
            EventRecord(
                event_type=event_type,
                path=path,
                when=now,
                source_path=source_path,
                dest_path=dest_path,
            )
        )
        self._increment_counter("events_enqueued")
        queue_size = self.event_queue.qsize()
        queue_warning_size = int(self.config.get("queue_warning_size", 500))
        if queue_size > queue_warning_size:
            self._increment_counter("queue_warning_count")
            self.log.warning(
                "Event queue size %s exceeds configured warning threshold %s",
                queue_size,
                queue_warning_size,
            )

    def classify_category(self, path: Path) -> str:
        text = str(path).lower()
        suffix = path.suffix.lower()
        rules = self.config.get("category_rules", {})
        # path keyword rules first
        for category, rule in rules.items():
            for keyword in rule.get("path_keywords", []):
                if keyword.lower() in text:
                    return category
        for category, rule in rules.items():
            if suffix in {ext.lower() for ext in rule.get("extensions", [])}:
                return category
        return "R&D"

    def classify_importance(self, path: Path) -> str:
        text = str(path).lower()
        rules = self.config.get("importance_keywords", {})
        for level in ["High", "Medium", "Low"]:
            for kw in rules.get(level, []):
                if kw.lower() in text:
                    return level
        return "Medium"

    def derive_parent_folder(self, path: Path) -> str:
        try:
            return str(path.parent.relative_to(Path(self.config["laptop_root"])))
        except Exception:
            return path.parent.name

    def safe_copy(self, source: Path, dest_root: Optional[str]) -> Dict[str, Optional[str]]:
        if not dest_root:
            self._increment_counter("copy_skipped")
            return {
                "status": "skipped_by_policy",
                "destination_path": None,
                "notes": "Copy destination is not configured.",
            }
        if not source.exists():
            self._increment_counter("copy_skipped")
            return {
                "status": "missing_source",
                "destination_path": None,
                "notes": "Source path does not exist.",
            }
        if not source.is_file():
            self._increment_counter("copy_skipped")
            return {
                "status": "not_a_file",
                "destination_path": None,
                "notes": "Source path is not a regular file.",
            }
        dest_root_path = Path(dest_root)
        if not dest_root_path.exists():
            self._increment_counter("copy_skipped")
            return {
                "status": "not_mounted",
                "destination_path": None,
                "notes": f"Destination root is not mounted: {dest_root_path}",
            }
        try:
            root = Path(self.config["laptop_root"])
            try:
                relative = source.relative_to(root)
            except Exception:
                relative = Path(source.name)
            target = dest_root_path / relative
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, target)
            self._increment_counter("copy_success")
            return {
                "status": "ok",
                "destination_path": str(target),
                "notes": "",
            }
        except Exception as exc:
            self.log.exception("Copy failed for %s -> %s", source, dest_root)
            self._increment_counter("copy_failed")
            return {
                "status": f"failed:{exc.__class__.__name__}",
                "destination_path": None,
                "notes": str(exc),
            }

    def decide_action(self, record: EventRecord) -> Dict[str, Optional[str]]:
        path = record.path
        when = record.when
        notes = []

        if record.event_type == "deleted":
            self._increment_counter("copy_skipped")
            return {
                "action": "deleted_logged",
                "backup_status": "not_needed",
                "destination_path": None,
                "notes": "Deleted path logged; copy skipped.",
            }

        if record.event_type == "moved":
            if record.source_path:
                notes.append(f"Moved from: {record.source_path}")
            if record.dest_path:
                notes.append(f"Moved to: {record.dest_path}")
            if not path.exists():
                notes.append("Destination path missing; copy skipped.")
                self._increment_counter("copy_skipped")
                return {
                    "action": "moved_logged",
                    "backup_status": "missing_source",
                    "destination_path": None,
                    "notes": " | ".join(notes),
                }

        if not path.exists():
            self._increment_counter("copy_skipped")
            return {
                "action": "kept_on_laptop",
                "backup_status": "missing_source",
                "destination_path": None,
                "notes": "Source path does not exist; copy skipped.",
            }

        if not path.is_file():
            self._increment_counter("copy_skipped")
            return {
                "action": "kept_on_laptop",
                "backup_status": "not_a_file",
                "destination_path": None,
                "notes": "Path is not a regular file; copy skipped.",
            }

        action = "kept_on_laptop"
        backup_status = "skipped_by_policy"
        destination_path = None
        copy_attempted = False

        if self.config.get("copy_recent_to_microsd", True):
            copy_attempted = True
            microsd_result = self.safe_copy(path, self.config.get("microsd_root"))
            microsd_status = microsd_result["status"]
            action = "copied_to_microsd" if microsd_status == "ok" else action
            backup_status = microsd_status or "failed:Unknown"
            if microsd_result.get("destination_path"):
                destination_path = microsd_result["destination_path"]
                notes.append(f"Copied to microSD: {destination_path}")
            if microsd_result.get("notes"):
                notes.append(f"microSD: {microsd_result['notes']}")

        stable_days = int(self.config.get("stable_days_before_wd", 7))
        try:
            mtime = datetime.fromtimestamp(path.stat().st_mtime)
            is_stable = (when - mtime).days >= stable_days
        except FileNotFoundError:
            is_stable = False

        if self.config.get("copy_stable_to_wd", True) and is_stable:
            copy_attempted = True
            wd_result = self.safe_copy(path, self.config.get("wd_root"))
            wd_status = wd_result["status"]
            if wd_status == "ok":
                action = "archived_to_wd"
                backup_status = "ok"
                destination_path = wd_result["destination_path"]
                notes.append(f"Copied to WD: {destination_path}")
            elif backup_status == "skipped_by_policy":
                backup_status = wd_status or "failed:Unknown"
            if wd_result.get("notes"):
                notes.append(f"WD: {wd_result['notes']}")

        if not copy_attempted:
            self._increment_counter("copy_skipped")

        return {
            "action": action,
            "backup_status": backup_status,
            "destination_path": destination_path,
            "notes": " | ".join(notes),
        }

    @staticmethod
    def _format_timestamp(value: datetime) -> str:
        return value.isoformat()

    def _build_event_payload(
        self,
        *,
        event_id: str,
        record: EventRecord,
        category: str,
        importance: str,
        parent_folder: str,
        handled_action: str,
        backup_status: str,
        session_key: Optional[str],
        notes: str,
    ) -> Dict[str, Any]:
        return {
            "event_id": event_id,
            "event_type": record.event_type,
            "event_time": self._format_timestamp(record.when),
            "file_name": record.path.name,
            "full_path": str(record.path),
            "extension": record.path.suffix.lower(),
            "parent_folder": parent_folder,
            "category": category,
            "importance": importance,
            "device": self.config.get("device", "LAP-002"),
            "source": "PythonWatcher",
            "session_key": session_key,
            "handled_action": handled_action,
            "backup_status": backup_status,
            "notes": notes,
            "project": self._derive_project(parent_folder),
        }

    @staticmethod
    def _derive_project(parent_folder: str) -> str:
        cleaned = (parent_folder or "").strip()
        if not cleaned or cleaned == ".":
            return ""
        return cleaned.replace("\\", "/").split("/")[0]

    def _write_sqlite_event(self, event_payload: Dict[str, Any]) -> None:
        try:
            insert_event(self.database_path, event_payload)
        except Exception:
            self.log.error(
                "Failed to write SQLite event for %s",
                event_payload.get("event_id"),
                exc_info=True,
            )

    def _write_sqlite_session(
        self,
        *,
        session_key: str,
        record: EventRecord,
        category: str,
        parent_folder: str,
        notes: str,
    ) -> None:
        try:
            insert_session(
                self.database_path,
                {
                    "session_key": session_key,
                    "start_time": self._format_timestamp(record.when),
                    "end_time": self._format_timestamp(record.when),
                    "duration_minutes": 1,
                    "category": category,
                    "project": self._derive_project(parent_folder),
                    "primary_file": record.path.name,
                    "event_count": 1,
                    "device": self.config.get("device", "LAP-002"),
                    "source": "PythonWatcher",
                    "notes": notes,
                },
            )
        except Exception:
            self.log.error(
                "Failed to write SQLite session for %s",
                session_key,
                exc_info=True,
            )

    def _write_fallback_event(
        self,
        *,
        event_payload: Dict[str, Any],
        failed_stage: str,
        error: Exception,
    ) -> bool:
        fallback_record = {
            "fallback_time": self._format_timestamp(datetime.now()),
            **event_payload,
            "error_type": error.__class__.__name__,
            "error_message": str(error),
            "failed_stage": failed_stage,
        }
        try:
            self.fallback_writer.append_record(fallback_record)
            self._increment_counter("fallback_events_written")
            return True
        except Exception:
            self._increment_counter("fallback_events_failed")
            self.log.critical(
                "Failed to write fallback event for %s during %s",
                event_payload["event_id"],
                failed_stage,
                exc_info=True,
            )
            return False

    def process_record(self, record: EventRecord) -> None:
        path = record.path
        if not path.exists() and record.event_type not in {"moved", "deleted"}:
            return

        event_id = f"EVT-{self.event_counter:06d}"
        try:
            category = self.classify_category(path)
            importance = self.classify_importance(path)
            parent_folder = self.derive_parent_folder(path)
            decision = self.decide_action(record)
            notes = decision.get("notes") or ""

            try:
                session_key = self.logger.append_or_extend_session(
                    {
                        "start_time": record.when,
                        "end_time": record.when,
                        "category": category,
                        "primary_path": parent_folder,
                        "owner": self.config.get("owner", "Karne"),
                    },
                    gap_minutes=int(self.config.get("session_gap_minutes", 10)),
                )
            except Exception as exc:
                fallback_written = self._write_fallback_event(
                    event_payload=self._build_event_payload(
                        event_id=event_id,
                        record=record,
                        category=category,
                        importance=importance,
                        parent_folder=parent_folder,
                        handled_action=decision["action"] or "kept_on_laptop",
                        backup_status=decision["backup_status"] or "failed:Unknown",
                        session_key=None,
                        notes=notes,
                    ),
                    failed_stage="session_write",
                    error=exc,
                )
                self._write_sqlite_event(
                    self._build_event_payload(
                        event_id=event_id,
                        record=record,
                        category=category,
                        importance=importance,
                        parent_folder=parent_folder,
                        handled_action=decision["action"] or "kept_on_laptop",
                        backup_status=decision["backup_status"] or "failed:Unknown",
                        session_key=None,
                        notes=notes,
                    )
                )
                self.log.error(
                    "Failed to write session for %s; event captured in fallback log=%s",
                    event_id,
                    fallback_written,
                    exc_info=True,
                )
                return

            self._write_sqlite_session(
                session_key=session_key,
                record=record,
                category=category,
                parent_folder=parent_folder,
                notes=notes,
            )

            event_payload = self._build_event_payload(
                event_id=event_id,
                record=record,
                category=category,
                importance=importance,
                parent_folder=parent_folder,
                handled_action=decision["action"] or "kept_on_laptop",
                backup_status=decision["backup_status"] or "failed:Unknown",
                session_key=session_key,
                notes=notes,
            )

            row = [
                event_id,
                record.when,
                record.event_type,
                path.name,
                str(path),
                path.suffix.lower(),
                parent_folder,
                category,
                importance,
                self.config.get("device", "LAP-002"),
                "PythonWatcher",
                session_key,
                decision["action"] or "kept_on_laptop",
                decision["backup_status"] or "failed:Unknown",
                notes,
            ]
            try:
                self.logger.append_activity(row)
                self._write_sqlite_event(event_payload)
                self.log.info("Logged %s | %s | %s", category, record.event_type, path)
            except Exception as exc:
                fallback_written = self._write_fallback_event(
                    event_payload=event_payload,
                    failed_stage="activity_write",
                    error=exc,
                )
                self._write_sqlite_event(event_payload)
                self.log.error(
                    "Failed to write activity for %s; event captured in fallback log=%s",
                    event_id,
                    fallback_written,
                    exc_info=True,
                )
        finally:
            self.event_counter += 1

    def worker_loop(self) -> None:
        while not self.stop_event.is_set() or not self.event_queue.empty():
            try:
                record = self.event_queue.get(timeout=1)
            except queue.Empty:
                continue
            try:
                self.process_record(record)
                self._increment_counter("events_processed")
            except PermissionError:
                self._increment_counter("events_failed")
                self.log.warning("Permission error on %s", record.path)
            except Exception:
                self._increment_counter("events_failed")
                self.log.exception("Failed to process %s", record.path)
            finally:
                self.event_queue.task_done()

    def start(self) -> None:
        self.validate_runtime()
        handler = DebouncedHandler(self)
        for watch_path in self.validated_paths["watch_paths"]:
            self.observer.schedule(handler, str(watch_path), recursive=True)
            self.log.info("Watching %s", watch_path)

        self.worker_thread = threading.Thread(
            target=self.worker_loop,
            name="XROIQWorkIntelligenceWorker",
        )
        self.worker_thread.start()
        try:
            self.observer.start()
            self.observer_started = True
            self.log.info("%s started", APP_NAME)
            while not self.stop_event.is_set():
                self._log_heartbeat_if_due()
                time.sleep(1)
        except KeyboardInterrupt:
            self.log.info("Stopping on keyboard interrupt")
        finally:
            self.stop()

    def stop(self) -> None:
        with self.stop_lock:
            if self.shutdown_started:
                return
            self.shutdown_started = True

        self.stop_event.set()
        if self.observer_started:
            self.observer.stop()
            self.observer.join(timeout=10)
            self.observer_started = False

        drain_timeout = int(self.config.get("shutdown_drain_timeout_seconds", 10))
        if not self._wait_for_queue_drain(drain_timeout):
            self.log.warning(
                "Queue drain timed out after %s seconds; unfinished_tasks=%s queue_size=%s",
                drain_timeout,
                self.event_queue.unfinished_tasks,
                self.event_queue.qsize(),
            )

        if (
            self.worker_thread is not None
            and self.worker_thread.is_alive()
            and threading.current_thread() is not self.worker_thread
        ):
            self.worker_thread.join(timeout=drain_timeout)
            if self.worker_thread.is_alive():
                self.log.warning(
                    "Worker thread did not stop within %s seconds",
                    drain_timeout,
                )

        self._log_heartbeat_if_due(force=True)
        self.log.info("Stopped | %s", self._format_counter_summary())


if __name__ == "__main__":
    config_path_was_explicit = len(sys.argv) > 1
    cfg = Path(sys.argv[1]) if config_path_was_explicit else DEFAULT_CONFIG_PATH
    service = WorkIntelligenceService(
        cfg,
        config_path_was_explicit=config_path_was_explicit,
    )
    service.start()
